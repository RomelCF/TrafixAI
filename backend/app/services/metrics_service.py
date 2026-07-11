# -*- coding: utf-8 -*-
"""
backend/app/services/metrics_service.py
Servicio para la generación de reportes y exportación de métricas a Excel.
"""
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.app.services.simulation_service import obtener_runner

def generar_excel_reporte(localidad: str) -> tuple[bytes, str]:
    """Genera un archivo Excel (.xlsx) con el reporte de comparación de escenarios y estado global."""
    r = obtener_runner(localidad)
    m = r.obtener_metricas()
    comparativa = m.get("comparativa", {})
    globales    = m.get("globales", {})
    sin = comparativa.get("sin_ia", {})
    con = comparativa.get("con_ia", {})

    nombre_localidad = "Centro Historico de Cusco" if localidad == "centro_historico" else "Sector Angostura"
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
    fase_label = {
        0: "Sin ejecutar", 1: "Fase 1 en curso (Sin IA)",
        2: "Fase 2 en curso (Con IA)", 3: "Completado",
    }.get(comparativa.get("benchmark_fase", 0), "Sin ejecutar")

    def fmt_mejora(val_sin, val_con, lower_better=True):
        try:
            v_sin, v_con = float(val_sin), float(val_con)
        except (TypeError, ValueError):
            return "Sin datos"
        if v_sin == 0:
            return "N/A"
        pct = 100*(v_sin-v_con)/v_sin if lower_better else 100*(v_con-v_sin)/v_sin
        return f"{'+' if pct >= 0 else ''}{pct:.1f}%"

    fill_gris_oscuro = PatternFill(fill_type="solid", fgColor="595959")
    fill_gris_medio  = PatternFill(fill_type="solid", fgColor="A6A6A6")
    fill_gris_claro  = PatternFill(fill_type="solid", fgColor="D9D9D9")
    fill_gris_fila   = PatternFill(fill_type="solid", fgColor="F2F2F2")

    ft_titulo  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ft_seccion = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ft_header  = Font(name="Calibri", bold=True, size=10, color="000000")
    ft_bold    = Font(name="Calibri", bold=True, size=10, color="000000")
    ft_normal  = Font(name="Calibri", size=10, color="000000")

    thin   = Side(style="thin",   color="000000")
    medium = Side(style="medium", color="000000")
    borde_datos    = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    borde_cabecera = Border(left=medium, right=medium, top=medium, bottom=medium)

    al_center = Alignment(horizontal="center", vertical="center")
    al_left   = Alignment(horizontal="left",   vertical="center")
    al_right  = Alignment(horizontal="right",  vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 22

    def fila_merged(row, texto, font, fill, align, altura=18):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=texto)
        c.font = font; c.fill = fill; c.alignment = align
        ws.row_dimensions[row].height = altura

    def set_cell(row, col, valor, font, align, fill=None, border=None):
        c = ws.cell(row=row, column=col, value=valor)
        c.font = font; c.alignment = align
        if fill:   c.fill   = fill
        if border: c.border = border
        return c

    f = 1

    fila_merged(f, "REPORTE DE COMPARACION DE ESCENARIOS DE TRAFICO",
                ft_titulo, fill_gris_oscuro, al_center, altura=24)
    f += 1
    f += 1

    meta = [
        ("Localidad",        nombre_localidad),
        ("Hora Simulada",    m.get("hora_simulada", "N/A") + "h"),
        ("Generado el",      ahora),
        ("IA activa",        "Si" if m.get("ia_activa") else "No"),
        ("Estado Benchmark", fase_label),
    ]
    for label, valor in meta:
        set_cell(f, 1, label, ft_bold, al_left)
        ws.merge_cells(start_row=f, start_column=2, end_row=f, end_column=4)
        set_cell(f, 2, valor, ft_normal, al_left)
        ws.row_dimensions[f].height = 15
        f += 1

    f += 1

    fila_merged(f, "COMPARATIVA DE ESCENARIOS", ft_seccion, fill_gris_medio, al_center, altura=18)
    f += 1

    cab = ["Metrica", "Escenario 1 - Tiempo Fijo (Sin IA)",
           "Escenario 2 - Adaptativo (Con IA)", "Mejora (+ = IA gana)"]
    for col_i, h in enumerate(cab, 1):
        set_cell(f, col_i, h, ft_header, al_center, fill_gris_claro, borde_cabecera)
    ws.row_dimensions[f].height = 16
    f += 1

    comp_rows = [
        ("Congestion Promedio",
         f"{sin.get('congestiva',0):.1f}%",   f"{con.get('congestiva',0):.1f}%",
         fmt_mejora(sin.get("congestiva"), con.get("congestiva"), True)),
        ("Velocidad Promedio",
         f"{sin.get('velocidad',0):.2f} px/frame", f"{con.get('velocidad',0):.2f} px/frame",
         fmt_mejora(sin.get("velocidad"), con.get("velocidad"), False)),
        ("Vehiculos Detenidos (Prom.)",
         f"{sin.get('detenidos',0):.1f} veh",  f"{con.get('detenidos',0):.1f} veh",
         fmt_mejora(sin.get("detenidos"), con.get("detenidos"), True)),
        ("Espera Promedio por Vehiculo",
         f"{sin.get('espera',0):.1f} min",     f"{con.get('espera',0):.1f} min",
         fmt_mejora(sin.get("espera"), con.get("espera"), True)),
        ("Flujo Vehicular Completado",
         f"{sin.get('flujo',0):.1f} veh/h",   f"{con.get('flujo',0):.1f} veh/h",
         fmt_mejora(sin.get("flujo"), con.get("flujo"), False)),
        ("Efectividad de Semaforos",
         f"{sin.get('efectividad',0):.1f}%",   f"{con.get('efectividad',0):.1f}%",
         fmt_mejora(sin.get("efectividad"), con.get("efectividad"), False)),
        ("Vehiculos Desviados por IA",
         str(sin.get("desviados",0)),          str(con.get("desviados",0)),
         f"+{con.get('desviados',0)} redirigidos"),
        ("Muestras (frames) analizadas",
         str(sin.get("frames",0)),             str(con.get("frames",0)), "N/A"),
    ]
    for i, (met, v_sin, v_con, mejora) in enumerate(comp_rows):
        fill_row = fill_gris_fila if i % 2 == 0 else None
        set_cell(f, 1, met,   ft_normal, al_left,  fill_row, borde_datos)
        set_cell(f, 2, v_sin, ft_normal, al_right, fill_row, borde_datos)
        set_cell(f, 3, v_con, ft_normal, al_right, fill_row, borde_datos)
        set_cell(f, 4, mejora,ft_normal, al_center,fill_row, borde_datos)
        ws.row_dimensions[f].height = 14
        f += 1

    f += 1

    fila_merged(f, "ESTADO ACTUAL DE LA SIMULACION", ft_seccion, fill_gris_medio, al_center, altura=18)
    f += 1

    for col_i, h in enumerate(["Metrica Global", "Valor Actual", "", ""], 1):
        set_cell(f, col_i, h, ft_header, al_center, fill_gris_claro, borde_cabecera)
    ws.row_dimensions[f].height = 16
    f += 1

    estado_rows = [
        ("Congestion Global",       f"{globales.get('congestion_global',0):.1f}%"),
        ("Velocidad Promedio",      f"{globales.get('velocidad_promedio',0):.2f} px/frame"),
        ("Vehiculos Activos",       str(globales.get("vehiculos_activos",0))),
        ("Vehiculos Detenidos",     str(globales.get("detenidos",0))),
        ("Espera Promedio",         f"{globales.get('espera_promedio',0):.1f} min"),
        ("Efectividad Semaforos",   f"{globales.get('efectividad_semaforos',0):.1f}%"),
        ("Emisiones CO2 Estimadas", f"{globales.get('emisiones_co2',0):.1f} g/km"),
        ("Tiempo Medio de Viaje",   f"{globales.get('tiempo_medio_viaje',0):.1f} min"),
    ]
    for i, (label, valor) in enumerate(estado_rows):
        fill_row = fill_gris_fila if i % 2 == 0 else None
        set_cell(f, 1, label, ft_normal, al_left,  fill_row, borde_datos)
        set_cell(f, 2, valor, ft_normal, al_right, fill_row, borde_datos)
        set_cell(f, 3, "",    ft_normal, al_left,  fill_row, borde_datos)
        set_cell(f, 4, "",    ft_normal, al_left,  fill_row, borde_datos)
        ws.row_dimensions[f].height = 14
        f += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()
    
    filename = f"reporte_trafico_{localidad}.xlsx"
    return data, filename
